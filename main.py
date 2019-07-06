import handlename
import re
import os
from openpyxl import load_workbook
import glob
import datetime


from handlename import PictureFileName, no_vietnamese

# t = PictureFileName("698 TRẦN VẠN NHIỆM.jpg", 0)




def handleName(name, methodType, rowStart, method, execute = False, excludeFirstC = False):
    rootSource = '../result/'
    # function handle change name if matched.
    # def handleChangeName():
    #     try:
    #         os.rename(o.filePath, rootSource + name + '/' + str(cell_cmnd) + '.jpg')
    #         text = 'OK: ' + cell_name, ' is changed to ', cell_cmnd
    #         print(text)
    #         aListOutput.append(text)
    #     except FileExistsError:
    #         os.rename(o.filePath,
    #                   rootSource + name + '/' + str(cell_cmnd) + '_' + cell_name + '.jpg')
    #         text = 'Warning: ' + cell_name, ' is changed to ', str(
    #             cell_cmnd) + '_' + cell_name + ' due to duplicated identify number.'
    #         print(text)
    #         aListOutput.append(text)
    #     except FileNotFoundError:
    #         text = 'Error: not found picture file ', o.filePath
    #         print(text)
    #         aListOutput.append(text)

    # get list file name.
    filePath1 = rootSource + name + '/*.jpg'
    filePath2 = rootSource + name + '/*.JPG'
    dataPath = rootSource + name + '/' + name + '.xlsx'
    aListFile = [f for f_ in [glob.glob(e) for e in [filePath1]] for f in f_] #list picture file name
    # for e in aListFile:
    #     try:
    #         os.rename(e, rootSource + name + '/' + '-'.join(e.split('-')[-2:]).strip())
    #     except FileNotFoundError:
    #         print(e)
    # return
    aListOutput = [] #using to trace log

    aListFileName = []
    aListDuplicate = []
    aRawName = []
    for item in aListFile:
        fileName = item[len(rootSource + name) + 1:]
        o = PictureFileName(item,fileName, methodType, excludeFirstC)
        o.filePath = rootSource + name + '/' + o.name
        aListFileName.append(o)

        # This code use for checking data analyze!!!
        # try:
        #     print(item, ' ', o.no_accent_name, o.date)
        # except AttributeError:
        #     print('oop!')

    for o in aListFileName:
        if o.raw_name != 'jpg':
            if o.raw_name not in aRawName:
                aRawName.append(o.raw_name)
            else:
                aListDuplicate.append(o.raw_name)
    # for i in aListFileName:
    #     print(i.no_accent_name, i.STT, i.date)
    # return

    wb = load_workbook(dataPath)
    # select demo.xlsx
    # from openpyxl import load_workbook
    # wb2 = load_workbook('test.xlsx')
    # ws4 = wb2["New Title"]
    sheet = wb.active
    # get max row count
    max_row = sheet.max_row
    # get max column count
    max_column = sheet.max_column
    # return
    # check column index.
    aListExcel = []
    for i in range(rowStart, max_row + 1):
        if sheet.cell(row=i, column=1) and sheet.cell(row=i, column=1).value and isinstance(
                sheet.cell(row=i, column=1).value, int) and sheet.cell(row=i, column=1).value >= 1:
            info = ''
            cell_stt = sheet.cell(row=i, column=1).value
            cell_name = sheet.cell(row=i, column=2).value
            if cell_name:
                if sheet.cell(row=i, column=4).value:
                    cell_namSinh = sheet.cell(row=i, column=4).value
                else:
                    cell_namSinh = sheet.cell(row=i, column=3).value
                cell_cmnd = sheet.cell(row=i, column=6).value
                cell_tinh = sheet.cell(row=i, column=5).value
                cell_fileName = sheet['L' + str(i)].internal_value
                cell_execute = sheet['M' + str(i)].internal_value
                cell_item = sheet['N' + str(i)].internal_value
        else:
            continue
        aListExcel.append({
            'index': i,
            'stt': cell_stt,
            'name': cell_name,
            'date': cell_namSinh,
            'id': cell_cmnd,
            'province': cell_tinh,
            'fileName': cell_fileName,
            'execute': cell_execute,
            'item': cell_item
        })
        # print(cell_name)
    print(aListExcel)
    a = 'skip'
    # return

    def handleChangeName(filePath, id, fileName):
        try:
            os.rename(filePath, rootSource + name + '/' + str(id) + '.jpg')
            text = 'OK: ' + fileName, ' is changed to ', id
            print(text)
            aListOutput.append(text)
        except FileExistsError:
            os.rename(filePath,
                      rootSource + name + '/' + str(id) + '_' + fileName + '.jpg')
            text = 'Warning: ' + fileName, ' is changed to ', str(
                id) + '_' + fileName + ' due to duplicated identify number.'
            print(text)
            aListOutput.append(text)
        except FileNotFoundError:
            text = 'Error: not found picture file ', filePath
            print(text)
            aListOutput.append(text)

    if execute:
        for oExcel in aListExcel:
            if oExcel['execute'] == 'True':
                handleChangeName(oExcel['item'], oExcel['id'], oExcel['fileName'])
                writeListToTextFile(aListOutput, rootSource + name + '/log.txt')
    else:
        for oImg in aListFileName:
            for oExcel in aListExcel:
                if method == 'Default':
                    if oImg.raw_name == oExcel['name'].lower():
                        if oImg.raw_name in aListDuplicate:
                            print(oImg.name)
                        else:
                            oExcel['fileName'] = oImg.name
                            sheet['L' + str(oExcel['index'])] = oImg.name
                            sheet['M' + str(oExcel['index'])] = 'True'
                            sheet['N' + str(oExcel['index'])] = oImg.item
                if method == 'STT':
                    if oImg.STT == oExcel['stt']:
                        oExcel['fileName'] = oImg.name
                        sheet['L' + str(oExcel['index'])] = oImg.name
                        sheet['M' + str(oExcel['index'])] = 'True'
                        sheet['N' + str(oExcel['index'])] = oImg.item

        wb.save(dataPath)





def writeListToTextFile(list, filePath, mode='a'):
    ''' Write list to csv line by line '''
    with open(filePath, mode, encoding="utf8") as myfile:
        for item in list:
            myfile.write(str(item) +  '\n')

# handleName('BV_DaKhoaTinhBinhPhuoc', 0, 8)
# handleName('TT_YTeHuyenLocNinh', 1, 6)
# handleName('TT_YTeHuyenHonQuan', 1, 9)
# handleName('BV_HoanMyBinhPhuoc', 2, 9)
# handleName('BV_YHocCoTruyenBinhPhuoc', 2, 7)
# handleName('CC_DanSoBinhPhuoc', 2, 12)
# handleName('CC_VeSinhAnToanThucPham', 2, 15)
# handleName('CT_CoPhanDuocVatTuYTeDopharco', 2, 8)
# handleName('TT_GiamDinhPhapYTinhBinhPhuoc', 2, 6)
# handleName('TT_KiemNghiemDuocPhamMyPhamTinhBinhPhuoc', 2, 8)
# handleName('TT_KiemSoatBenhTatTinhBinhPhuoc', 2, 7)
# handleName('TT_YTeBuGiaMap', 2, 8)
# handleName('TT_YTeHuyenDongPhu', 2, 8)
# handleName('TT_YTeHuyenLocNinh', 2, 6)
# handleName('TT_YTeThiXaBinhLong', 2, 10)
# handleName('TT_YTeThiXaDongXoai', 2, 5)


# part2

# handleName('TrungTamYTeHuyenChonThanh', 2, 9)
# handleName('TrungTamYTeHuyenPhuRieng', 1, 8)
# handleName('TruongCaoDangYTeBinhPhuoc', 2, 9)

# handleName('BuDang', 2, 12)
# handleName('BuDop', 2, 8)
# handleName('BVTinh', 4, 8, True)
# handleName('TrungTam', 2, 13)

# handleName('HonQuan', 4, 9, method='STT', execute=True)
# handleName('TTYTDongXoai', 4, 12, method='Default', execute=True, excludeFirstC = True)
# handleName('YHCT', 4, 11, method='Default', execute=True, excludeFirstC = False)
# handleName('BuGiaMap', 4, 10, method='Default', execute=True, excludeFirstC = False)
# handleName('BinhLong', 4, 13, method='STT', execute=True, excludeFirstC = False)
# handleName('HoanMy', 4, 9, method='Default', execute=True, excludeFirstC = False)
handleName('PhuocLong', 4, 13, method='Default', execute=True, excludeFirstC = False)

